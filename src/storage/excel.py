"""Excel storage implementation for crypto call data."""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError(
        "openpyxl is required for Excel storage. Install with: pip install openpyxl"
    )

logger = logging.getLogger(__name__)


class ExcelStorage:
    """Excel-based storage implementation for crypto call data.

    This class provides persistent storage using Excel (.xlsx) files with proper
    error handling and resource management.
    """

    def __init__(self, file_path: Path) -> None:
        """Initialize Excel storage with file path.

        Args:
            file_path: Path to the Excel file.

        Raises:
            Exception: If Excel file initialization fails.
        """
        self.file_path = file_path
        self._workbook: Optional[Workbook] = None
        self._worksheet: Optional[Worksheet] = None
        self._header_map: Dict[str, int] = {}
        self._is_closed = False
        self._init_workbook()

    def _init_workbook(self) -> None:
        """Initialize workbook and worksheet, create if doesn't exist.

        Raises:
            Exception: If workbook initialization fails.
        """
        try:
            if self.file_path.exists():
                # Load existing workbook
                self._workbook = load_workbook(self.file_path)
                if "crypto_calls" in self._workbook.sheetnames:
                    self._worksheet = self._workbook["crypto_calls"]
                else:
                    # Create the sheet if it doesn't exist
                    self._worksheet = self._workbook.create_sheet("crypto_calls")
                    self._create_headers()
            else:
                # Create new workbook
                self._workbook = Workbook()
                # Remove default sheet and create our sheet
                if "Sheet" in self._workbook.sheetnames:
                    self._workbook.remove(self._workbook["Sheet"])
                self._worksheet = self._workbook.create_sheet("crypto_calls")
                self._create_headers()
                self._save_workbook()

            self._read_headers()  # Always read headers after init
            logger.info(f"Excel workbook initialized at {self.file_path}")

        except Exception as e:
            logger.error(f"Failed to initialize Excel workbook: {e}")
            raise

    def _read_headers(self) -> None:
        """Read header row and create a map of header names to column indices."""
        if not self._worksheet:
            raise Exception("Worksheet is not available")

        self._header_map = {}
        for col, cell in enumerate(self._worksheet[1], 1):
            if cell.value:
                self._header_map[str(cell.value)] = col

        if not self._header_map:
            logger.warning("Excel sheet has no headers. Re-creating them.")
            self._create_headers()
            # Reread headers after creating them
            for col, cell in enumerate(self._worksheet[1], 1):
                if cell.value:
                    self._header_map[str(cell.value)] = col

        logger.debug(f"Excel headers mapped: {self._header_map}")

    def _create_headers(self) -> None:
        """Create column headers in the worksheet."""
        if not self._worksheet:
            raise Exception("Worksheet is not available")

        headers = [
            "token_name",
            "entry_cap",
            "peak_cap",
            "x_gain",
            "vip_x",
            "message_type",
            "contract_address",
            "time_to_peak",
            "linked_crypto_call_id",
            "timestamp",
            "message_id",
            "channel_name",
        ]

        # Check if we need to create headers
        needs_headers = False
        if self._worksheet.max_row == 0:
            # Completely empty sheet
            needs_headers = True
        elif self._worksheet.max_row == 1:
            # Check if first row is empty or has different headers
            first_row_empty = all(cell.value is None for cell in self._worksheet[1])
            if first_row_empty:
                needs_headers = True

        if needs_headers:
            for col, header in enumerate(headers, 1):
                self._worksheet.cell(row=1, column=col, value=header)
            self._header_map = {header: i for i, header in enumerate(headers, 1)}
            logger.debug("Created Excel headers")

    def _save_workbook(self) -> None:
        """Save the workbook to file.

        Raises:
            Exception: If save operation fails.
        """
        if not self._workbook:
            raise Exception("Workbook is not available")

        try:
            self._workbook.save(self.file_path)
            logger.debug(f"Excel workbook saved to {self.file_path}")
        except Exception as e:
            logger.error(f"Failed to save Excel workbook: {e}")
            raise

    def append_row(self, data: Dict[str, Any]) -> None:
        """Append a new row of crypto call data to the Excel file.

        Args:
            data: Dictionary containing crypto call data to store.
                 Expected keys: token_name, entry_cap, peak_cap, x_gain,
                 vip_x, message_type, contract_address, time_to_peak,
                 linked_crypto_call_id, timestamp, message_id, channel_name

        Raises:
            ValueError: If data is None or not a dictionary.
            Exception: If Excel operation fails.
        """
        if self._is_closed:
            raise Exception("Excel storage is closed")

        if data is None:
            raise ValueError("Data cannot be None")

        if not isinstance(data, dict):
            raise TypeError("Data must be a dictionary")

        if not data:
            raise ValueError("Data dictionary cannot be empty")

        if not self._worksheet or not self._workbook:
            raise Exception("Excel worksheet is not available")

        if not self._header_map:
            raise Exception("Excel headers not loaded. Cannot append row.")

        try:
            # Find the next empty row
            next_row = self._worksheet.max_row + 1

            # Insert data into cells based on header map
            for key, value in data.items():
                if key in self._header_map:
                    col = self._header_map[key]
                    self._worksheet.cell(row=next_row, column=col, value=value)
                else:
                    logger.warning(
                        f"Column '{key}' not found in Excel headers. Skipping."
                    )

            # Save the workbook
            self._save_workbook()
            logger.debug(
                f"Inserted crypto call data for token: {data.get('token_name')} (type: {data.get('message_type')})"
            )

        except Exception as e:
            logger.error(f"Failed to insert data into Excel: {e}")
            raise

    def get_records(self, limit: Optional[int] = None) -> List[Dict[str, Any]]:
        """Retrieve records from the Excel file.

        Args:
            limit: Maximum number of records to return. If None, returns all records.

        Returns:
            List of dictionaries containing stored crypto call data.

        Raises:
            Exception: If Excel operation fails.
        """
        if self._is_closed:
            raise Exception("Excel storage is closed")

        if not self._worksheet or not self._workbook:
            raise Exception("Excel worksheet is not available")

        try:
            records = []

            # Get headers from first row using the stored map
            if not self._header_map:
                self._read_headers()  # Ensure headers are read

            headers = sorted(self._header_map.keys(), key=lambda h: self._header_map[h])

            if not headers:
                return []

            # Read data rows (starting from row 2)
            max_row = self._worksheet.max_row
            if max_row <= 1:
                return []

            # Determine how many records to read
            end_row = max_row
            if limit is not None:
                start_row = max(2, max_row - limit + 1)
            else:
                start_row = 2

            for row_num in range(start_row, max_row + 1):
                row_data = {}
                for header, col_idx in self._header_map.items():
                    cell_value = self._worksheet.cell(row=row_num, column=col_idx).value
                    row_data[header] = cell_value
                records.append(row_data)

            logger.debug(f"Retrieved {len(records)} records from Excel")
            return records

        except Exception as e:
            logger.error(f"Failed to retrieve records from Excel: {e}")
            raise

    def close(self) -> None:
        """Close Excel workbook and cleanup resources.

        This method should be called when finished with the storage instance
        to properly clean up Excel resources.
        """
        if not self._is_closed:
            try:
                if self._workbook:
                    # Save one final time before closing
                    self._save_workbook()
                    self._workbook.close()
                    self._workbook = None
                    self._worksheet = None
                self._is_closed = True
                logger.debug("Excel workbook closed")
            except Exception as e:
                logger.error(f"Error closing Excel workbook: {e}")
                # Don't re-raise as this is cleanup code

    def get_crypto_call_by_message_id(self, message_id: int) -> Optional[int]:
        """Get crypto call ID by message ID - not implemented for Excel storage."""
        logger.warning(
            "get_crypto_call_by_message_id not implemented for Excel storage"
        )
        return None

    def get_crypto_call_by_id(self, call_id: int) -> Optional[Dict[str, Any]]:
        """Get crypto call by ID - not implemented for Excel storage."""
        logger.warning("get_crypto_call_by_id not implemented for Excel storage")
        return None

    def find_related_discovery(
        self,
        channel_name: str,
        token_name: Optional[str] = None,
        contract_address: Optional[str] = None,
        entry_cap: Optional[float] = None,
        since_hours: int = 24,
    ) -> Optional[int]:
        """Find related discovery call - not implemented for Excel storage."""
        logger.warning("find_related_discovery not implemented for Excel storage")
        return None
